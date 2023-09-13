from django.shortcuts import render
from django.views.generic import TemplateView, ListView, DetailView, CreateView, UpdateView, DeleteView, FormView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.utils import timezone
import openpyxl
from datetime import datetime
from django.db.models import Q, FilteredRelation

from master.models import Cost, Affiliation, Employee, AccountingPeriod
from master.forms import CostForm, CostSearchForm, CostImportForm

class CostListView(LoginRequiredMixin, ListView):
    template_name = 'master/cost/index.html'
    model = Cost
    paginate_by = 10
    context_object_name = 'items'

    def nvl(self, value, default=''):
        if value is None:
            val = default
        else:
            val = str(value).strip()
        return val

    def convert_string_to_date(self, value, defalut=None):
        
        val = self.nvl(value, defalut)

        if val is not None:
            date_val = datetime.strptime(str(val), '%Y-%m-%d %H:%M:%S')
            val = date_val.strftime('%Y-%m-%d')

        return val

    def get_queryset(self):

        # param = 'init' の場合 セッションクリア
        if self.request.GET.get('param', '') == 'init':
            if 'cost_search_values' in self.request.session:
                del self.request.session['cost_search_values']

        # sessionに値がある場合、その値でクエリ発行する。
        if 'cost_search_values' in self.request.session:

            # セッションから検索情報取得
            search_values = self.request.session['cost_search_values']

            # 検索条件
            condition1 = Q()
            condition2 = Q()
            condition3 = Q()
            condition4 = Q()
            condition5 = Q()

            # 検索文字が指定されている場合、条件設定
            if len(search_values[0]) != 0:
                condition1 = Q(affiliation__accounting_period_id=search_values[0])
            else:
                condition1 = Q(affiliation__gte=0)

            # 検索文字が指定されている場合、条件設定
            if len(search_values[1]) != 0:
                condition2 = Q(affiliation__employee__name__icontains=search_values[1])
                condition3 = Q(affiliation__employee__employee_no__icontains=search_values[1])
                condition4 = Q(affiliation__employee__email__icontains=search_values[1])
                condition5 = Q(affiliation__department__department_name__icontains=search_values[1])

            # 検索条件を指定して検索結果取得
            result = Cost.objects.select_related('affiliation')\
                    .filter(condition1, condition2 | condition3 | condition4 | condition5)

        else:
            # 検索条件指定なしの場合、全件取得
            result = Cost.objects.select_related('affiliation')

        result = result\
                    .order_by(\
                        '-affiliation__accounting_period__accounting_period'\
                        ,'affiliation__department__display_order'\
                        ,'affiliation__position__display_order'\
                        ,'affiliation__grade__display_order')\
                    .values(\
                         'affiliation__accounting_period__accounting_period'\
                        ,'affiliation__department__department_name'\
                        ,'affiliation__position__position_name'\
                        ,'affiliation__grade__grade_name'\
                        ,'affiliation__employee__name'\
                        ,'pk','cost1', 'cost2', 'cost3', 'cost4', 'cost_total', 'updated', 'created')

        print(result.query)
        return result
    
    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)

        # sessionに値がある場合、その値をセットする。（ページングしてもform値が変わらないように）
        if 'cost_search_values' in self.request.session:
            search_values = self.request.session['cost_search_values']
            default_data = {
                'search_accounting_period': search_values[0], 
                'keyword': search_values[1],
            }
        else:
            default_data = {
                'search_accounting_period': '', 
                'keyword': '',
            }

        context['search_form'] = CostSearchForm(initial=default_data)
        context['import_form'] = CostImportForm()

        return context

    def post(self, request, *args, **kwargs):

        try:
            # アップロード時
            if self.request.POST.get('mode', '') == 'upload':

                upload_file = self.request.FILES['upload_file']
                accounting_period_id = self.request.POST.get('search_accounting_period', '')

                # 役職ファイルアプロード処理
                if not self.import_cost(upload_file, accounting_period_id):
                    print('import_cost Error !!')

            # 検索時
            elif self.request.POST.get('mode', '') == 'search':

                # セッションに検索値を設定
                search_values = [
                    self.request.POST.get('search_accounting_period', ''),
                    self.request.POST.get('keyword', ''),
                ]
                request.session['cost_search_values'] = search_values

        except Exception as e:
            print(e)

        finally:
            # 検索時にページネーションに関連したエラーを防ぐ
            self.request.GET = self.request.GET.copy()
            self.request.GET.clear()
            return self.get(request, *args, **kwargs)

    def import_cost(self, upload_file, accounting_period_id):

        try:
            # アップロードファイルのオープン
            wb = openpyxl.load_workbook(upload_file, data_only=True)

            # 先頭シートを参照
            ws = wb.worksheets[0]

            # 一括追加用配列
            insert_items = []
            update_items = []

            # 会計期データ取得
            lists = AccountingPeriod.objects.filter(id=accounting_period_id)

            if len(lists) == 0:
                print('会計期を登録してください。')
                accounting_period = None
            else:
                accounting_period = lists[0]

                i = 2
                while i <= ws.max_row:

                    # b列（社員番号）の取得
                    col_b = str(ws.cell(row=i, column=2).value)

                    # 所属データ取得
                    lists = Affiliation.objects.select_related('accounting_period','employee').\
                                filter(accounting_period_id=accounting_period.id, employee__employee_no=str(col_b))
                    print(lists.query)

                    if len(lists) == 0:
                        print('会計期：' + str(accounting_period_id) + ' 社員番号：' + col_b + ' の所属を登録してください。')
                        affiliation = None
                    else:
                        affiliation = lists[0]

                        # 所属IDが同じデータ取得
                        result = Cost.objects.filter(
                            affiliation_id=int(affiliation.id))

                        # 同じデータ存在確認
                        if not result.exists():

                            # 存在しない場合、新規登録
                            item = Cost(
                                affiliation = affiliation,
                                cost1 = int(ws.cell(row=i, column=4).value),
                                cost2 = int(ws.cell(row=i, column=5).value),
                                cost3 = int(ws.cell(row=i, column=7).value),
                                cost4 = int(ws.cell(row=i, column=8).value),
                                cost_total = int(ws.cell(row=i, column=9).value),
                            )

                            insert_items.append(item)
                        else:
                            # 存在する場合、更新
                            item = result.get()
                            
                            cost1 = int(ws.cell(row=i, column=4).value),
                            cost2 = int(ws.cell(row=i, column=5).value),
                            cost3 = int(ws.cell(row=i, column=7).value),
                            cost4 = int(ws.cell(row=i, column=8).value),
                            cost_total = int(ws.cell(row=i, column=9).value),
                            item.updated = datetime.now()

                            update_items.append(item)

                    i = i + 1

                # 更新リストデータが存在する場合
                if len(update_items) > 0:
                    Cost.objects.bulk_update(update_items, ['cost1','cost2','cost3','cost4','cost_total','updated'])

                # 新規リストデータが存在する場合
                if len(insert_items) > 0:
                    Cost.objects.bulk_create(insert_items)

            wb.close()
            rc = True

        except Exception as e:
            print(e)
            rc = False

        return rc

class CostCreateView(LoginRequiredMixin, CreateView):
    template_name = 'master/cost/create.html'
    model = Cost
    form_class = CostForm
    success_url = reverse_lazy('master.cost.index')

class CostUpdateView(LoginRequiredMixin, UpdateView):
    template_name = 'master/cost/update.html'
    model = Cost
    form_class = CostForm
    success_url = reverse_lazy('master.cost.index')
    context_object_name = 'item'

    def form_valid(self, form):
        Cost = form.save(commit=False)
        Cost.updated = timezone.now()
        Cost.save()
        return super().form_valid(form)

class CostDeleteView(LoginRequiredMixin, DeleteView):
    template_name = 'master/cost/delete.html'
    model = Cost
    success_url = reverse_lazy('master.cost.index')
    context_object_name = 'item'
