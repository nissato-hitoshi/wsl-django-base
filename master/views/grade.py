from django.shortcuts import render
from django.views.generic import TemplateView, ListView, DetailView, CreateView, UpdateView, DeleteView, FormView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.utils import timezone
import openpyxl
from datetime import datetime
from django.db.models import Q

from master.models import Grade
from master.forms import GradeForm, GradeSearchForm, GradeImportForm

class GradeListView(LoginRequiredMixin, ListView):
    template_name = 'master/grade/index.html'
    model = Grade
    paginate_by = 10
    context_object_name = 'items'

    def get_queryset(self):

        # param = 'init' の場合 セッションクリア
        if self.request.GET.get('param', '') == 'init':
            if 'grade_search_values' in self.request.session:
                del self.request.session['grade_search_values']

        # sessionに値がある場合、その値でクエリ発行する。
        if 'grade_search_values' in self.request.session:

            # セッションから検索情報取得
            search_values = self.request.session['grade_search_values']

            # 検索条件
            condition1 = Q()
            condition2 = Q()

            # 検索文字が指定されている場合、条件設定
            if len(search_values[0]) != 0:
                condition1 = Q(grade_name__icontains=search_values[0])
                condition2 = Q(grade_code__icontains=search_values[0])

            # 検索条件を指定して検索結果取得
            return Grade.objects.select_related().filter(condition1 | condition2).order_by('display_order')

        else:
            # 検索条件指定なしの場合、全件取得
            return Grade.objects.all().order_by('display_order')

    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)

        # sessionに値がある場合、その値をセットする。（ページングしてもform値が変わらないように）
        if 'grade_search_values' in self.request.session:
            search_values = self.request.session['grade_search_values']
            default_data = {
                'keyword': search_values[0],
            }
        else:
            default_data = {
                'keyword': '',
            }
        
        context['search_form'] = GradeSearchForm(initial=default_data)
        context['import_form'] = GradeImportForm()

        return context

    def post(self, request, *args, **kwargs):

        try:
            # アップロード時
            if self.request.POST.get('mode', '') == 'upload':

                upload_file = self.request.FILES['upload_file']

                # 資格ファイルアプロード処理
                if not self.import_grade(upload_file):
                    print('import_grade Error !!')

            # 検索時
            elif self.request.POST.get('mode', '') == 'search':

                # セッションに検索値を設定
                search_values = [
                    self.request.POST.get('keyword', ''),
                ]
                request.session['grade_search_values'] = search_values

        except Exception as e:
            print(e)

        finally:
            # 検索時にページネーションに関連したエラーを防ぐ
            self.request.GET = self.request.GET.copy()
            self.request.GET.clear()
            return self.get(request, *args, **kwargs)

    def import_grade(self, upload_file):
        try:
            # アップロードファイルのオープン
            wb = openpyxl.load_workbook(upload_file, data_only=True)

            # 先頭シートを参照
            ws = wb.worksheets[0]

            # 一括追加用配列
            insert_items = []
            update_items = []

            i = 2
            while i <= ws.max_row:

                # 同じ資格コードのデータ取得
                result = Grade.objects.filter(grade_code=str(ws.cell(row=i, column=1).value))

                # 同じ資格コードデータ存在確認
                if not result.exists():

                    # 存在しない場合、新規登録
                    item = Grade(
                        grade_code = ws.cell(row=i, column=1).value,
                        grade_name = ws.cell(row=i, column=2).value,
                        display_order = ws.cell(row=i, column=3).value,
                    )

                    insert_items.append(item)
                else:
                    # 存在する場合、更新
                    item = result.get()
                    
                    item.grade_name = ws.cell(row=i, column=2).value
                    item.display_order = ws.cell(row=i, column=3).value
                    item.updated = datetime.now()

                    update_items.append(item)

                i = i + 1

            # 更新リストデータが存在する場合
            if len(update_items) > 0:
                Grade.objects.bulk_update(update_items, ['grade_name', 'display_order','updated'])

            # 新規リストデータが存在する場合
            if len(insert_items) > 0:
                Grade.objects.bulk_create(insert_items)

            wb.close()
            rc = True

        except Exception as e:
            print(e)
            rc = False

        return rc

class GradeCreateView(LoginRequiredMixin, CreateView):
    template_name = 'master/grade/create.html'
    model = Grade
    form_class = GradeForm
    success_url = reverse_lazy('master.grade.index')

class GradeUpdateView(LoginRequiredMixin, UpdateView):
    template_name = 'master/grade/update.html'
    model = Grade
    form_class = GradeForm
    success_url = reverse_lazy('master.grade.index')
    context_object_name = 'item'

    def form_valid(self, form):
        Grade = form.save(commit=False)
        Grade.updated = timezone.now()
        Grade.save()
        return super().form_valid(form)

class GradeDeleteView(LoginRequiredMixin, DeleteView):
    template_name = 'master/grade/delete.html'
    model = Grade
    success_url = reverse_lazy('master.grade.index')
    context_object_name = 'item'
