from django.shortcuts import render
from django.views.generic import TemplateView, ListView, DetailView, CreateView, UpdateView, DeleteView, FormView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.utils import timezone
import openpyxl
from datetime import datetime
from django.db.models import Q

from master.models import Position
from master.forms import PositionForm, PositionSearchForm, PositionImportForm
from . import BaseView

class PositionListView(LoginRequiredMixin, ListView, BaseView):
    template_name = 'master/position/index.html'
    model = Position
    paginate_by = 10
    context_object_name = 'items'

    def get_queryset(self):

        # param = 'init' の場合 セッションクリア
        if self.request.GET.get('param', '') == 'init':
            if 'position_search_values' in self.request.session:
                del self.request.session['position_search_values']

        # sessionに値がある場合、その値でクエリ発行する。
        if 'position_search_values' in self.request.session:

            # セッションから検索情報取得
            search_values = self.request.session['position_search_values']

            # 検索条件
            condition1 = Q()
            condition2 = Q()

            # 検索文字が指定されている場合、条件設定
            if len(search_values[0]) != 0:
                condition1 = Q(position_name__icontains=search_values[0])
                condition2 = Q(position_code__icontains=search_values[0])

            # 検索条件を指定して検索結果取得
            return Position.objects.select_related().filter(condition1 | condition2).order_by('display_order')

        else:
            # 検索条件指定なしの場合、全件取得
            return Position.objects.all().order_by('display_order')

    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)

        # sessionに値がある場合、その値をセットする。（ページングしてもform値が変わらないように）
        if 'position_search_values' in self.request.session:
            search_values = self.request.session['position_search_values']
            default_data = {
                'keyword': search_values[0],
            }
        else:
            default_data = {
                'keyword': '',
            }
        
        context['search_form'] = PositionSearchForm(initial=default_data)
        context['import_form'] = PositionImportForm()

        return context

    def post(self, request, *args, **kwargs):

        try:
            # アップロード時
            if self.request.POST.get('mode', '') == 'upload':

                upload_file = self.request.FILES['upload_file']

                # 役職ファイルアプロード処理
                if not self.import_position(upload_file):
                    print('import_position Error !!')

            # 検索時
            elif self.request.POST.get('mode', '') == 'search':

                # セッションに検索値を設定
                search_values = [
                    self.request.POST.get('keyword', ''),
                ]
                request.session['position_search_values'] = search_values

        except Exception as e:
            print(e)

        finally:
            # 検索時にページネーションに関連したエラーを防ぐ
            self.request.GET = self.request.GET.copy()
            self.request.GET.clear()
            return self.get(request, *args, **kwargs)

    def import_position(self, upload_file):
        try:
            # アップロードファイルのオープン
            wb = openpyxl.load_workbook(upload_file, data_only=True)

            # 先頭シートを参照
            ws = wb.worksheets[0]

            # 一括追加用配列
            insert_items = []
            update_items = []

            i = 2
            while i <= ws.max_row:

                # 同じ役職コードのデータ取得
                result = Position.objects.filter(position_code=str(ws.cell(row=i, column=1).value))

                # 同じ役職コードデータ存在確認
                if not result.exists():

                    # 存在しない場合、新規登録
                    item = Position(
                        position_code = ws.cell(row=i, column=1).value,
                        position_name = ws.cell(row=i, column=2).value,
                        display_order = ws.cell(row=i, column=3).value,
                    )

                    insert_items.append(item)
                else:
                    # 存在する場合、更新
                    item = result.get()
                    
                    item.position_name = ws.cell(row=i, column=2).value
                    item.display_order = ws.cell(row=i, column=3).value
                    item.updated = datetime.now()

                    update_items.append(item)

                i = i + 1

            # 更新リストデータが存在する場合
            if len(update_items) > 0:
                Position.objects.bulk_update(update_items, ['position_name', 'display_order','updated'])

            # 新規リストデータが存在する場合
            if len(insert_items) > 0:
                Position.objects.bulk_create(insert_items)

            wb.close()
            rc = True

        except Exception as e:
            print(e)
            rc = False

        return rc

class PositionCreateView(LoginRequiredMixin, CreateView):
    template_name = 'master/position/create.html'
    model = Position
    form_class = PositionForm
    success_url = reverse_lazy('master.position.index')

class PositionUpdateView(LoginRequiredMixin, UpdateView):
    template_name = 'master/position/update.html'
    model = Position
    form_class = PositionForm
    success_url = reverse_lazy('master.position.index')
    context_object_name = 'item'

    def form_valid(self, form):
        Position = form.save(commit=False)
        Position.updated = timezone.now()
        Position.save()
        return super().form_valid(form)

class PositionDeleteView(LoginRequiredMixin, DeleteView):
    template_name = 'master/position/delete.html'
    model = Position
    success_url = reverse_lazy('master.position.index')
    context_object_name = 'item'
