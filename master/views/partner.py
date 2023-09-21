from django.shortcuts import render
from django.views.generic import TemplateView, ListView, DetailView, CreateView, UpdateView, DeleteView, FormView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.utils import timezone
import openpyxl
from datetime import datetime
from django.db.models import Q

from master.models import Partner
from master.forms import PartnerForm, PartnerSearchForm, PartnerImportForm
from . import BaseView

class PartnerListView(LoginRequiredMixin, ListView, BaseView):
    template_name = 'master/partner/index.html'
    model = Partner
    paginate_by = 10
    context_object_name = 'items'

    def get_queryset(self):

        # param = 'init' の場合 セッションクリア
        if self.request.GET.get('param', '') == 'init':
            if 'partner_search_values' in self.request.session:
                del self.request.session['partner_search_values']

        # sessionに値がある場合、その値でクエリ発行する。
        if 'partner_search_values' in self.request.session:

            # セッションから検索情報取得
            search_values = self.request.session['partner_search_values']

            # 検索条件
            condition1 = Q()
            condition2 = Q()
            condition3 = Q()
            condition4 = Q()

            # 検索文字が指定されている場合、条件設定
            if len(search_values[0]) != 0:
                condition1 = Q(partner_name__icontains=search_values[0])
                condition2 = Q(supplier_name__icontains=search_values[0])
                condition3 = Q(description__icontains=search_values[0])
                condition4 = Q(partner_name_k__icontains=search_values[0])

            # 検索条件を指定して検索結果取得
            result = Partner.objects.all().filter(condition1 | condition2 | condition3 | condition4)
        else:
            # 検索条件指定なしの場合、全件取得
            result = Partner.objects.all()

        result = result.order_by('display_order', '-updated', 'id')

        #print(result.query)
        return result
    
    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)

        # sessionに値がある場合、その値をセットする。（ページングしてもform値が変わらないように）
        if 'partner_search_values' in self.request.session:
            search_values = self.request.session['partner_search_values']
            default_data = {
                'keyword': search_values[0],
            }
        else:
            default_data = {
                'keyword': '',
            }
        
        context['search_form'] = PartnerSearchForm(initial=default_data)
        context['import_form'] = PartnerImportForm()

        return context

    def post(self, request, *args, **kwargs):

        try:
            # アップロード時
            if self.request.POST.get('mode', '') == 'upload':

                upload_file = self.request.FILES['upload_file']

                # 役職ファイルアプロード処理
                if not self.import_partner(upload_file):
                    print('import_partner Error !!')

            # 検索時
            elif self.request.POST.get('mode', '') == 'search':

                # セッションに検索値を設定
                search_values = [
                    self.request.POST.get('keyword', ''),
                ]
                request.session['partner_search_values'] = search_values

        except Exception as e:
            print(e)

        finally:
            # 検索時にページネーションに関連したエラーを防ぐ
            self.request.GET = self.request.GET.copy()
            self.request.GET.clear()
            return self.get(request, *args, **kwargs)

    def import_partner(self, upload_file):
        try:
            # アップロードファイルのオープン
            wb = openpyxl.load_workbook(upload_file, data_only=True)

            # 先頭シートを参照
            ws = wb.worksheets[0]

            # 一括追加用配列
            insert_items = []
            update_items = []

            i = 2
            while i <= ws.max_row:

                # 同じ仕入先、パートナー名のデータ取得
                result = Partner.objects.filter(\
                         partner_name=str(ws.cell(row=i, column=3).value)\
                        ,supplier_name=str(ws.cell(row=i, column=1).value))

                # 同じ仕入先、パートナー名データ存在確認
                if not result.exists():

                    # 存在しない場合、新規登録
                    item = Partner(
                        supplier_name = ws.cell(row=i, column=1).value,
                        description = ws.cell(row=i, column=2).value,
                        partner_name = ws.cell(row=i, column=3).value,
                        partner_name_k = ws.cell(row=i, column=4).value,
                        display_order = ws.cell(row=i, column=5).value,
                    )

                    insert_items.append(item)
                else:
                    # 存在する場合、更新
                    item = result.get()
                    
                    item.description = ws.cell(row=i, column=2).value
                    item.partner_name_k = ws.cell(row=i, column=4).value
                    item.display_order = ws.cell(row=i, column=5).value
                    item.updated = datetime.now()

                    update_items.append(item)

                i = i + 1

            # 更新リストデータが存在する場合
            if len(update_items) > 0:
                Partner.objects.bulk_update(update_items, ['partner_name_k', 'description', 'display_order', 'updated'])

            # 新規リストデータが存在する場合
            if len(insert_items) > 0:
                Partner.objects.bulk_create(insert_items)

            wb.close()
            rc = True

        except Exception as e:
            print(e)
            rc = False

        return rc

class PartnerCreateView(LoginRequiredMixin, CreateView):
    template_name = 'master/partner/create.html'
    model = Partner
    form_class = PartnerForm
    success_url = reverse_lazy('master.partner.index')

class PartnerUpdateView(LoginRequiredMixin, UpdateView):
    template_name = 'master/partner/update.html'
    model = Partner
    form_class = PartnerForm
    success_url = reverse_lazy('master.partner.index')
    context_object_name = 'item'

    def form_valid(self, form):
        Department = form.save(commit=False)
        Department.updated = timezone.now()
        Department.save()
        return super().form_valid(form)

class PartnerDeleteView(LoginRequiredMixin, DeleteView):
    template_name = 'master/partner/delete.html'
    model = Partner
    success_url = reverse_lazy('master.partner.index')
    context_object_name = 'item'
