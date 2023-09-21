from django.shortcuts import render
from django.views.generic import TemplateView, ListView, DetailView, CreateView, UpdateView, DeleteView, FormView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.utils import timezone
import openpyxl
from datetime import datetime
from django.db.models import Q

from master.models import AccountingPeriod
from master.forms import AccountingPeriodForm, AccountingPeriodSearchForm, AccountingPeriodImportForm
from . import BaseView

class AccountingPeriodListView(LoginRequiredMixin, ListView, BaseView):
    template_name = 'master/accounting_period/index.html'
    model = AccountingPeriod
    paginate_by = 10
    context_object_name = 'items'

    def get_queryset(self):

        # param = 'init' の場合 セッションクリア
        if self.request.GET.get('param', '') == 'init':
            if 'accounting_period_search_values' in self.request.session:
                del self.request.session['accounting_period_search_values']

        # sessionに値がある場合、その値でクエリ発行する。
        if 'accounting_period_search_values' in self.request.session:

            # セッションから検索情報取得
            search_values = self.request.session['accounting_period_search_values']

            # 検索条件
            condition1 = Q()

            # 検索文字が指定されている場合、条件設定
            if len(search_values) != 0:
                condition1 = Q(accounting_period__icontains=search_values[0])

            # 検索条件を指定して検索結果取得
            return AccountingPeriod.objects.select_related().filter(condition1)

        else:
            # 検索条件指定なしの場合、全件取得
            return AccountingPeriod.objects.all()

    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)

        # sessionに値がある場合、その値をセットする。（ページングしてもform値が変わらないように）
        if 'accounting_period_search_values' in self.request.session:
            search_values = self.request.session['accounting_period_search_values']
            default_data = {
                'keyword': search_values[0],
            }
        else:
            default_data = {
                'keyword': '',
            }

        context['search_form'] = AccountingPeriodSearchForm(initial=default_data)
        context['import_form'] = AccountingPeriodImportForm()

        return context

    def post(self, request, *args, **kwargs):

        try:
            # アップロード時
            if self.request.POST.get('mode', '') == 'upload':

                upload_file = self.request.FILES['upload_file']

                # 役職ファイルアプロード処理
                if not self.import_accounting_period(upload_file):
                    print('import_accounting_period Error !!')

            # 検索時
            elif self.request.POST.get('mode', '') == 'search':

                # セッションに検索値を設定
                search_values = [
                    self.request.POST.get('keyword', ''),
                ]
                request.session['accounting_period_search_values'] = search_values

        except Exception as e:
            print(e)

        finally:
            # 検索時にページネーションに関連したエラーを防ぐ
            self.request.GET = self.request.GET.copy()
            self.request.GET.clear()
            return self.get(request, *args, **kwargs)

    def import_accounting_period(self, upload_file):

        try:
            # アップロードファイルのオープン
            wb = openpyxl.load_workbook(upload_file, data_only=True)

            # 先頭シートを参照
            ws = wb.worksheets[0]

            # 一括追加用配列
            insert_items = []
            update_items = []

            i = 2
            while i <= ws.max_row:

                # 会計期と氏名が同じデータ取得
                result = AccountingPeriod.objects.filter(accounting_period=int(ws.cell(row=i, column=1).value))

                # 同じデータ存在確認
                if not result.exists():

                    # 存在しない場合、新規登録
                    item = AccountingPeriod(
                        accounting_period = ws.cell(row=i, column=1).value,
                        start_date = self.convert_string_to_date(ws.cell(row=i, column=2).value, None),
                        end_date = self.convert_string_to_date(ws.cell(row=i, column=3).value, None),
                    )

                    insert_items.append(item)
                else:
                    # 存在する場合、更新
                    item = result.get()
                    
                    item.start_date = self.convert_string_to_date(ws.cell(row=i, column=2).value, None)
                    item.end_date = self.convert_string_to_date(ws.cell(row=i, column=3).value, None)
                    item.updated = datetime.now()

                    update_items.append(item)

                i = i + 1

            # 更新リストデータが存在する場合
            if len(update_items) > 0:
                AccountingPeriod.objects.bulk_update(update_items, ['start_date','end_date','updated'])

            # 新規リストデータが存在する場合
            if len(insert_items) > 0:
                AccountingPeriod.objects.bulk_create(insert_items)

            wb.close()
            rc = True

        except Exception as e:
            print(e)
            rc = False

        return rc

class AccountingPeriodCreateView(LoginRequiredMixin, CreateView):
    template_name = 'master/accounting_period/create.html'
    model = AccountingPeriod
    form_class = AccountingPeriodForm
    success_url = reverse_lazy('master.accounting_period.index')

class AccountingPeriodUpdateView(LoginRequiredMixin, UpdateView):
    template_name = 'master/accounting_period/update.html'
    model = AccountingPeriod
    form_class = AccountingPeriodForm
    success_url = reverse_lazy('master.accounting_period.index')
    context_object_name = 'item'

    def form_valid(self, form):
        Affiliation = form.save(commit=False)
        Affiliation.updated = timezone.now()
        Affiliation.save()
        return super().form_valid(form)

class AccountingPeriodDeleteView(LoginRequiredMixin, DeleteView):
    template_name = 'master/accounting_period/delete.html'
    model = AccountingPeriod
    success_url = reverse_lazy('master.accounting_period.index')
    context_object_name = 'item'
