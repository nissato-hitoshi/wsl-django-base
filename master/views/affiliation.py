from django.shortcuts import render
from django.views.generic import TemplateView, ListView, DetailView, CreateView, UpdateView, DeleteView, FormView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.utils import timezone
import openpyxl
from datetime import datetime
from django.db.models import Q

from master.models import Affiliation, Department, Grade, Position, Employee, AccountingPeriod
from master.forms import AffiliationForm, AffiliationSearchForm, AffiliationImportForm
from . import BaseView

class AffiliationListView(LoginRequiredMixin, ListView, BaseView):
    template_name = 'master/affiliation/index.html'
    model = Affiliation
    paginate_by = 10
    context_object_name = 'items'

    def get_queryset(self):

        # param = 'init' の場合 セッションクリア
        if self.request.GET.get('param', '') == 'init':
            if 'affiliation_search_values' in self.request.session:
                del self.request.session['affiliation_search_values']

        # sessionに値がある場合、その値でクエリ発行する。
        if 'affiliation_search_values' in self.request.session:

            # セッションから検索情報取得
            search_values = self.request.session['affiliation_search_values']

            # 検索条件
            condition1 = Q()
            condition2 = Q()
            condition3 = Q()
            condition4 = Q()
            condition5 = Q()

            # 検索文字が指定されている場合、条件設定
            if len(search_values[0]) != 0:
                condition1 = Q(accounting_period=search_values[0])
            else:
                condition1 = Q(accounting_period_id__gte=0)

            if len(search_values[1]) != 0:
                condition2 = Q(employee__name__icontains=search_values[1])
                condition3 = Q(department__department_name__icontains=search_values[1])
                condition4 = Q(position__position_name__icontains=search_values[1])
                condition5 = Q(grade__grade_name__icontains=search_values[1])

            # 検索条件を指定して検索結果取得
            return Affiliation.objects.select_related().filter(condition1, condition2 | condition3 | condition4 | condition5).order_by(
                    'accounting_period', 'department__display_order', 'position__display_order', 'grade__display_order')
        else:
            # 検索条件指定なしの場合、全件取得
            return Affiliation.objects.all().order_by(
                    'accounting_period', 'department__display_order', 'position__display_order', 'grade__display_order')

    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)

        # sessionに値がある場合、その値をセットする。（ページングしてもform値が変わらないように）
        if 'affiliation_search_values' in self.request.session:
            search_values = self.request.session['affiliation_search_values']
            default_data = {
                'search_accounting_period': search_values[0], 
                'keyword': search_values[1],
            }
        else:
            default_data = {
                'search_accounting_period': '', 
                'keyword': '',
            }

        context['search_form'] = AffiliationSearchForm(initial=default_data)
        context['import_form'] = AffiliationImportForm()

        return context


    def post(self, request, *args, **kwargs):

        try:
            # アップロード時
            if self.request.POST.get('mode', '') == 'upload':

                upload_file = self.request.FILES['upload_file']

                # 役職ファイルアプロード処理
                if not self.import_affiliation(upload_file):
                    print('import_affiliation Error !!')

            # 検索時
            elif self.request.POST.get('mode', '') == 'search':

                # セッションに検索値を設定
                search_values = [
                    self.request.POST.get('search_accounting_period', ''),
                    self.request.POST.get('keyword', ''),
                ]
                request.session['affiliation_search_values'] = search_values

        except Exception as e:
            print(e)

        finally:
            # 検索時にページネーションに関連したエラーを防ぐ
            self.request.GET = self.request.GET.copy()
            self.request.GET.clear()
            return self.get(request, *args, **kwargs)

    def import_affiliation(self, upload_file):

        try:
            # アップロードファイルのオープン
            wb = openpyxl.load_workbook(upload_file, data_only=True)

            # 先頭シートを参照
            ws = wb.worksheets[0]

            # 一括追加用配列
            insert_items = []
            update_items = []

            i = 2
            while i <= ws.max_row:

                # 会計期の取得
                lists = AccountingPeriod.objects.filter(accounting_period=int(ws.cell(row=i, column=1).value))
                if len(lists) == 0:
                    accounting_period = None
                else:
                    accounting_period = lists[0]

                # 部門の取得
                lists = Department.objects.filter(department_name=str(ws.cell(row=i, column=2).value))
                if len(lists) == 0:
                    department = None
                else:
                    department = lists[0]

                # 役職の取得
                lists = Position.objects.filter(position_name=str(ws.cell(row=i, column=3).value))
                if len(lists) == 0:
                    position = None
                else:
                    position = lists[0]

                # 資格の取得
                lists = Grade.objects.filter(grade_name=str(ws.cell(row=i, column=4).value))
                if len(lists) == 0:
                    grade = None
                else:
                    grade = lists[0]

                # 社員情報の取得
                lists = Employee.objects.filter(name=str(ws.cell(row=i, column=5).value))
                if len(lists) == 0:
                    employee = None
                else:
                    employee = lists[0]

                # 社員情報登録済の場合のみ追加・更新対象
                if employee is None:
                    print(str(i) + '行目 未登録ユーザー：' + str(ws.cell(row=i, column=5).value))
                elif department is None:
                    print(str(i) + '行目 未登録部門：' + str(ws.cell(row=i, column=2).value))
                elif position is None:
                    print(str(i) + '行目 未登録役職：' + str(ws.cell(row=i, column=3).value))
                elif grade is None:
                    print(str(i) + '行目 未登録資格：' + str(ws.cell(row=i, column=4).value))
                elif accounting_period is None:
                    print(str(i) + '行目 未登録会計期：' + str(ws.cell(row=i, column=1).value))
                else:
                    # 会計期と氏名が同じデータ取得
                    result = Affiliation.objects.filter(
                        accounting_period=accounting_period
                        ,employee_id=int(employee.id))

                    # 同じデータ存在確認
                    if not result.exists():

                        # 存在しない場合、新規登録
                        item = Affiliation(
                            accounting_period = accounting_period,
                            department = department,
                            position = position,
                            grade = grade,
                            employee = employee,
                        )

                        insert_items.append(item)
                    else:
                        # 存在する場合、更新
                        item = result.get()
                        
                        item.department = department
                        item.position = position
                        item.grade = grade
                        item.updated = datetime.now()

                        update_items.append(item)

                i = i + 1

            # 更新リストデータが存在する場合
            if len(update_items) > 0:
                Affiliation.objects.bulk_update(update_items, ['department','position','grade','updated'])

            # 新規リストデータが存在する場合
            if len(insert_items) > 0:
                Affiliation.objects.bulk_create(insert_items)

            wb.close()
            rc = True

        except Exception as e:
            print(e)
            rc = False

        return rc

class AffiliationCreateView(LoginRequiredMixin, CreateView):
    template_name = 'master/affiliation/create.html'
    model = Affiliation
    form_class = AffiliationForm
    success_url = reverse_lazy('master.affiliation.index')

class AffiliationUpdateView(LoginRequiredMixin, UpdateView):
    template_name = 'master/affiliation/update.html'
    model = Affiliation
    form_class = AffiliationForm
    success_url = reverse_lazy('master.affiliation.index')
    context_object_name = 'item'

    def form_valid(self, form):
        Affiliation = form.save(commit=False)
        Affiliation.updated = timezone.now()
        Affiliation.save()
        return super().form_valid(form)

class AffiliationDeleteView(LoginRequiredMixin, DeleteView):
    template_name = 'master/affiliation/delete.html'
    model = Affiliation
    success_url = reverse_lazy('master.affiliation.index')
    context_object_name = 'item'
