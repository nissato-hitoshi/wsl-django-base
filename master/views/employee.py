from django.shortcuts import render
from django.views.generic import TemplateView, ListView, DetailView, CreateView, UpdateView, DeleteView, FormView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.utils import timezone
import openpyxl
from datetime import datetime
from django.db.models import Q

from master.models import Employee
from master.forms import EmployeeForm, EmployeeSearchForm, EmployeeImportForm
from . import BaseView

class EmployeeListView(LoginRequiredMixin, ListView, BaseView):
    template_name = 'master/employee/index.html'
    model = Employee
    paginate_by = 10
    context_object_name = 'items'

    def get_queryset(self):

        # param = 'init' の場合 セッションクリア
        if self.request.GET.get('param', '') == 'init':
            if 'employee_search_values' in self.request.session:
                del self.request.session['employee_search_values']

        # sessionに値がある場合、その値でクエリ発行する。
        if 'employee_search_values' in self.request.session:

            # セッションから検索情報取得
            search_values = self.request.session['employee_search_values']

            # 検索条件
            condition1 = Q()
            condition2 = Q()
            condition3 = Q()

            # 検索文字が指定されている場合、条件設定
            if len(search_values[0]) != 0:
                condition1 = Q(employee_no__icontains=search_values[0])
                condition2 = Q(name__icontains=search_values[0])
                condition3 = Q(email__icontains=search_values[0])

            # 検索条件を指定して検索結果取得
            return Employee.objects.select_related().filter(
                    condition1 | condition2 | condition3).order_by('-updated')

        else:
            # 検索条件指定なしの場合、全件取得
            return Employee.objects.select_related().all().order_by('-updated')

    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)

        # sessionに値がある場合、その値をセットする。（ページングしてもform値が変わらないように）
        if 'employee_search_values' in self.request.session:
            search_values = self.request.session['employee_search_values']
            default_data = {
                'keyword': search_values[0],
            }
        else:
            default_data = {
                'keyword': '',
            }

        context['search_form'] = EmployeeSearchForm(initial=default_data)
        context['import_form'] = EmployeeImportForm()

        return context

    def post(self, request, *args, **kwargs):

        try:
            # アップロード時
            if self.request.POST.get('mode', '') == 'upload':

                upload_file = self.request.FILES['upload_file']

                # ファイルインポート処理
                if not self.import_employee(upload_file):
                    print('import_employee Error !!')

            # 検索時
            elif self.request.POST.get('mode', '') == 'search':
                # セッションに検索値を設定
                search_values = [
                    self.request.POST.get('keyword', ''),
                ]
                self.request.session['employee_search_values'] = search_values

        except Exception as e:
            print(e)

        finally:
            # 検索時にページネーションに関連したエラーを防ぐ
            self.request.GET = self.request.GET.copy()
            self.request.GET.clear()
            return self.get(request, *args, **kwargs)

    def import_employee(self, upload_file):

        try:
            rc = True

            # アップロードファイルのオープン
            wb = openpyxl.load_workbook(upload_file, data_only=True)

            # 先頭シートを参照
            ws = wb.worksheets[0]

            # 一括追加用配列
            insert_items = []
            update_items = []

            i = 2
            while i <= ws.max_row:

                # 同じ社員番号のデータ取得
                result = Employee.objects.filter(employee_no=str(ws.cell(row=i, column=1).value))

                # 同じ社員番号データ存在確認
                if not result.exists():

                    # 存在しない場合、新規登録
                    item = Employee(
                        employee_no = ws.cell(row=i, column=1).value,
                        name = ws.cell(row=i, column=2).value,
                        employee_code = ws.cell(row=i, column=3).value,
                        email = ws.cell(row=i, column=4).value,
                        date_of_birth = self.convert_string_to_date(ws.cell(row=i, column=5).value),
                        hire_date = self.convert_string_to_date(ws.cell(row=i, column=6).value),
                        retirement_date = self.convert_string_to_date(ws.cell(row=i, column=7).value),
                    )

                    insert_items.append(item)
                else:
                    # 存在する場合、更新
                    item = result.get()
                    
                    item.name = ws.cell(row=i, column=2).value
                    item.employee_code = ws.cell(row=i, column=3).value
                    item.email = ws.cell(row=i, column=4).value
                    item.date_of_birth = self.convert_string_to_date(ws.cell(row=i, column=5).value, None)
                    item.hire_date = self.convert_string_to_date(ws.cell(row=i, column=6).value, None)
                    item.retirement_date = self.convert_string_to_date(ws.cell(row=i, column=7).value, None)
                    item.updated = datetime.now()

                    update_items.append(item)

                i = i + 1

            # 更新リストデータが存在する場合
            if len(update_items) > 0:
                Employee.objects.bulk_update(update_items, ['name', 'employee_code', 'email', 'date_of_birth', 'hire_date', 'retirement_date','updated'])

            # 新規リストデータが存在する場合
            if len(insert_items) > 0:
                Employee.objects.bulk_create(insert_items)

            wb.close()
            rc = True

        except Exception as e:
            print(e)
            rc = False

        return rc

class EmployeeCreateView(LoginRequiredMixin, CreateView):
    template_name = 'master/employee/create.html'
    model = Employee
    form_class = EmployeeForm
    success_url = reverse_lazy('master.employee.index')

class EmployeeUpdateView(LoginRequiredMixin, UpdateView):
    template_name = 'master/employee/update.html'
    model = Employee
    form_class = EmployeeForm
    success_url = reverse_lazy('master.employee.index')
    context_object_name = 'item'

    def form_valid(self, form):
        Employee = form.save(commit=False)
        Employee.updated = timezone.now()
        Employee.save()
        return super().form_valid(form)

class EmployeeDeleteView(LoginRequiredMixin, DeleteView):
    template_name = 'master/employee/delete.html'
    model = Employee
    success_url = reverse_lazy('master.employee.index')
    context_object_name = 'item'

